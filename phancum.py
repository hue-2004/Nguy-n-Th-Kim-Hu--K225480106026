import pandas as pd
from sklearn.cluster import KMeans
from sklearn.impute import SimpleImputer
import openpyxl

# =========================
# 1. Đường dẫn file Excel
# =========================
duong_dan_file = r"E:\KHDL\BTVN\diem.xlsx"

# =========================
# 2. Đọc dữ liệu Excel
# =========================
du_lieu = pd.read_excel(duong_dan_file, header=None)

# =========================
# 3. Hàm xử lý lỗi ngày tháng
# =========================
def xu_ly_ngay_thang(x):
    if isinstance(x, str) and x.startswith('2026-'):
        parts = x.split('-')
        if len(parts) == 3:
            return float(f"{int(parts[2])}.{int(parts[1])}")
    return x

# =========================
# 4. Lấy MSSV, tên SV, môn học
# =========================
danh_sach_mssv = du_lieu.iloc[1, 3:].values
danh_sach_ten = du_lieu.iloc[2, 3:].values
danh_sach_mon_hoc = du_lieu.iloc[4:, 2].values

# =========================
# 5. Lấy bảng điểm
# =========================
bang_diem = du_lieu.iloc[4:, 3:]

# Xử lý lỗi ngày tháng
bang_diem = bang_diem.map(xu_ly_ngay_thang)

# Chuyển sang số
bang_diem = bang_diem.apply(pd.to_numeric, errors='coerce')

# =========================
# 6. Chuyển vị ma trận
# =========================
bang_diem_chuyen_vi = bang_diem.T

# =========================
# 7. Điền dữ liệu khuyết
# =========================
imputer = SimpleImputer(strategy='mean')
bang_diem_da_dien = imputer.fit_transform(bang_diem_chuyen_vi)

# =========================
# 8. Phân cụm KMeans
# =========================
kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)

nhom_goc = kmeans.fit_predict(bang_diem_da_dien)

# =========================
# 9. Tính điểm trung bình
# =========================
diem_tb = bang_diem_da_dien.mean(axis=1)

# =========================
# 10. Sắp xếp cụm theo học lực
# =========================
bang_tam = pd.DataFrame({
    'Cum': nhom_goc,
    'DiemTB': diem_tb
})

diem_tb_theo_cum = bang_tam.groupby('Cum')['DiemTB'].mean()

# Sắp xếp giảm dần
cum_sap_xep = diem_tb_theo_cum.sort_values(ascending=False).index

# Ánh xạ cụm
anh_xa = {
    cum_sap_xep[0]: "Giỏi",
    cum_sap_xep[1]: "Khá",
    cum_sap_xep[2]: "Trung bình"
}

nhom_text = [anh_xa[x] for x in nhom_goc]

# =========================
# 11. Tạo bảng kết quả
# =========================
bang_ket_qua = pd.DataFrame(
    bang_diem_da_dien,
    columns=danh_sach_mon_hoc
)

bang_ket_qua.insert(0, 'MSSV', danh_sach_mssv)
bang_ket_qua.insert(1, 'Tên Sinh Viên', danh_sach_ten)

# Thêm điểm trung bình
bang_ket_qua['Điểm TB'] = diem_tb.round(1)

# Thêm nhóm KMeans
bang_ket_qua['Nhóm KMeans'] = nhom_text

# =========================
# 12. Xuất Excel + Ghi chú
# =========================
file_ket_qua = "Ket_qua_KMeans.xlsx"

with pd.ExcelWriter(file_ket_qua, engine='openpyxl') as writer:
    
    # Ghi bảng kết quả
    bang_ket_qua.to_excel(writer, index=False, sheet_name='Ket_qua')
    
    # Lấy sheet
    sheet = writer.sheets['Ket_qua']
    
    # Xác định dòng cuối
    dong_cuoi = len(bang_ket_qua) + 1
    
    # =========================
    # Thêm phần ghi chú
    # =========================
    sheet.cell(row=dong_cuoi + 2, column=1, value="Ghi chú:")
    sheet.cell(row=dong_cuoi + 3, column=1, value="0: Giỏi nhất")
    sheet.cell(row=dong_cuoi + 4, column=1, value="1: Khá")
    sheet.cell(row=dong_cuoi + 5, column=1, value="2: Thấp nhất")

print("Đã chạy xong!")
print("Kết quả đã lưu tại:", file_ket_qua)